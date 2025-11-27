import pandas as pd
import os
import time
import sys
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# ---------------------------------------------------------
# 基础设置
# ---------------------------------------------------------
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
except:
    pass

print("===========================================")
print("      智能对账系统 V7.0 (Excel彩色增强版)")
print("===========================================")
print(f"当前工作目录: {os.getcwd()}\n")

# ---------------------------------------------------------
# 清洗工具
# ---------------------------------------------------------
def clean_text(text):
    if pd.isna(text): return ""
    s = str(text)
    # 去除 BOM, 幽灵空格, 全角空格, 普通空格
    s = s.replace('\ufeff', '').replace('\xa0', '').replace('\u3000', '').replace(' ', '')
    return s.strip()

def read_csv_smart(filename):
    if not os.path.exists(filename): return None
    encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']
    for enc in encodings:
        try:
            df = pd.read_csv(filename, encoding=enc)
            df.columns = [clean_text(c) for c in df.columns]
            return df
        except:
            continue
    return None

def main():
    print("1. 正在读取文件...")
    df_bang = read_csv_smart('磅单列表.csv')
    df_men = read_csv_smart('门禁数据.csv')

    if df_bang is None or df_men is None:
        print("❌ 错误：未找到 '磅单列表.csv' 或 '门禁数据.csv'")
        input("按回车键退出...")
        return

    # 配置列名
    K_BANG_PLATE = '车号'
    K_BANG_NAME_D = '存货名称'
    K_BANG_SPEC_E = '规格型号'
    K_BANG_WEIGHT = '净重'
    K_BANG_TIME = '毛重过磅时间'
    
    K_MEN_PLATE = '车牌号'
    K_MEN_NAME = '运输货物名称'
    K_MEN_WEIGHT = '运输货物净重'
    K_MEN_TIME = '出厂时间'

    # 校验列名
    missing = []
    if K_BANG_TIME not in df_bang.columns: missing.append(f"磅单缺: {K_BANG_TIME}")
    if K_MEN_TIME not in df_men.columns: missing.append(f"门禁缺: {K_MEN_TIME}")
    
    if missing:
        print(f"❌ 列名错误: {missing}")
        input("按回车键退出...")
        return

    print("2. 正在进行全表清洗与计算...")

    # 全表清洗
    def clean_df_content(df):
        obj_cols = df.select_dtypes(include=['object']).columns
        for col in obj_cols:
            df[col] = df[col].apply(lambda x: str(x).replace('\xa0', ' ').strip() if pd.notnull(x) else x)
        return df

    df_bang = clean_df_content(df_bang)
    df_men = clean_df_content(df_men)

    # 格式转换
    df_bang['dt_base'] = pd.to_datetime(df_bang[K_BANG_TIME], errors='coerce')
    df_men['dt_base'] = pd.to_datetime(df_men[K_MEN_TIME], errors='coerce')
    
    df_bang['num_weight'] = pd.to_numeric(df_bang[K_BANG_WEIGHT], errors='coerce').fillna(0)
    df_men['num_weight'] = pd.to_numeric(df_men[K_MEN_WEIGHT], errors='coerce').fillna(0)
    
    df_bang['key_plate'] = df_bang[K_BANG_PLATE].astype(str).str.replace(' ', '').str.upper()
    df_men['key_plate'] = df_men[K_MEN_PLATE].astype(str).str.replace(' ', '').str.upper()

    # 智能货名
    def get_real_name(row):
        d = str(row.get(K_BANG_NAME_D, '')).strip()
        e = str(row.get(K_BANG_SPEC_E, '')).strip()
        if not d or d.lower() == 'nan' or d.replace('.', '').isdigit():
            return e
        return d
    df_bang['有效货名'] = df_bang.apply(get_real_name, axis=1)

    # 核心比对
    results = []
    valid_men = df_men.dropna(subset=['dt_base'])

    for index, row in df_bang.iterrows():
        plate = row['key_plate']
        time_bang = row['dt_base']
        weight_bang = row['num_weight']
        name_bang = row['有效货名']
        
        # 默认结果
        res = {
            '|': '|', # 分隔符
            '门禁_匹配车牌': '未找到',
            '门禁_出厂时间': '',
            '门禁_净重': 0,
            '门禁_货物名称': '',
            '停留时长(分)': '',
            '结果_车牌': '🔴异常',
            '结果_净重': '🔴异常',
            '结果_货名': '🔴异常',
            '备注': ''
        }

        if pd.notnull(time_bang):
            matches = valid_men[valid_men['key_plate'] == plate]
            if not matches.empty:
                future = matches[matches['dt_base'] >= time_bang].copy()
                if not future.empty:
                    future['diff'] = future['dt_base'] - time_bang
                    best = future.sort_values('diff').iloc[0]
                    diff_min = best['diff'].total_seconds() / 60
                    
                    res['门禁_匹配车牌'] = best[K_MEN_PLATE]
                    res['门禁_出厂时间'] = best[K_MEN_TIME]
                    res['门禁_净重'] = best['num_weight']
                    res['门禁_货物名称'] = best[K_MEN_NAME]
                    res['停留时长(分)'] = round(diff_min, 1)
                    
                    # 判定
                    if diff_min > 2880:
                        res['结果_车牌'] = '🟡时间过长'
                        res['备注'] = '>48小时'
                    else:
                        res['结果_车牌'] = '🟢正常'
                    
                    if abs(weight_bang - best['num_weight']) <= 0.02:
                        res['结果_净重'] = '🟢正常'
                    else:
                        res['结果_净重'] = '🟡不符'
                        
                    m_name = str(best[K_MEN_NAME])
                    p_name = str(name_bang)
                    kws = ['焦', '煤', '油', '酸', '碱', '盐', '苯']
                    if m_name == p_name or m_name in p_name or p_name in m_name:
                        res['结果_货名'] = '🟢正常'
                    elif any(k in m_name and k in p_name for k in kws):
                        res['结果_货名'] = '🟢正常'
                    else:
                        res['结果_货名'] = '🔴不符'
                        res['备注'] = f"磅[{p_name}] vs 门[{m_name}]"
                else:
                    res['备注'] = '未出厂'
        
        row_data = row.to_dict()
        # 清理
        for k in ['dt_base', 'num_weight', 'key_plate', '有效货名']:
            if k in row_data: del row_data[k]
        row_data.update(res)
        results.append(row_data)

    # ---------------------------------------------------------
    # 导出 Excel 并上色
    # ---------------------------------------------------------
    print("3. 计算完成，正在生成彩色 Excel 报表...")
    
    df_final = pd.DataFrame(results)
    
    # 调整列顺序：原数据在前，结果在后
    res_cols = list(res.keys())
    org_cols = [c for c in df_final.columns if c not in res_cols]
    df_final = df_final[org_cols + res_cols]

    timestamp = time.strftime("%H点%M分%S秒")
    output_filename = f'对账结果_{timestamp}.xlsx'
    
    try:
        # 1. 写入 Excel
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='比对结果')
        
        # 2. 打开 Excel 进行美化
        wb = load_workbook(output_filename)
        ws = wb.active
        
        # 定义样式
        # 标题行样式
        fill_header_org = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # 灰色
        fill_header_res = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # 淡蓝
        
        # 内容区域样式
        fill_res_col = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid") # 极淡蓝(AliceBlue)
        
        font_red = Font(color="FF0000", bold=True)      # 红字粗体
        font_green = Font(color="008000", bold=True)    # 绿字粗体
        font_orange = Font(color="FF8C00", bold=True)   # 橙字粗体
        
        # 找到结果列的起始位置（通过分隔符 '|'）
        res_start_idx = 0
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == '|':
                res_start_idx = idx
                break
        
        max_row = ws.max_row
        max_col = ws.max_column

        # --- 循环处理单元格 ---
        for col in range(1, max_col + 1):
            col_letter = ws.cell(1, col).column_letter
            header_val = ws.cell(1, col).value
            
            # A. 设置表头颜色
            if col >= res_start_idx:
                ws.cell(1, col).fill = fill_header_res
            else:
                ws.cell(1, col).fill = fill_header_org
                
            # B. 设置列宽自适应 (简化版)
            ws.column_dimensions[col_letter].width = 15

            # C. 设置结果列的背景色 和 文字颜色
            if col >= res_start_idx:
                for row in range(2, max_row + 1):
                    cell = ws.cell(row, col)
                    # 1. 背景色：淡蓝色，区分新数据
                    cell.fill = fill_res_col
                    
                    # 2. 文字颜色：根据内容变色
                    val = str(cell.value)
                    if "异常" in val or "不符" in val:
                        cell.font = font_red
                    elif "正常" in val:
                        cell.font = font_green
                    elif "过长" in val:
                        cell.font = font_orange

        wb.save(output_filename)
        print(f"\n✅✅✅ 成功！已生成彩色报表: [{output_filename}]")
        
    except Exception as e:
        print(f"\n❌ Excel 生成失败: {e}")
        try:
            # 降级方案：如果Excel生成失败，存CSV
            csv_name = f'对账结果_备份_{timestamp}.csv'
            df_final.to_csv(csv_name, index=False, encoding='utf-8-sig')
            print(f"已自动降级保存为 CSV: {csv_name}")
        except:
            pass

    print("\n(按回车键关闭)")
    input()

if __name__ == '__main__':
    main()
